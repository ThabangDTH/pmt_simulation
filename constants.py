from openpyxl.styles import Font, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Application Constants
# ---------------------------------------------------------------------------

MISSING_VALUE = -1

INPUT_SHEET_HOUSEHOLDS = "Households"
INPUT_SHEET_MEMBERS = "Members"
OUTPUT_SHEET_RESULTS = "PMT_Results"
OUTPUT_SHEET_SUMMARY = "Summary"

# ---------------------------------------------------------------------------
# PMT coefficients
# ---------------------------------------------------------------------------

INTERCEPT = 5.318

COEFFICIENTS = {
    "size2": -0.469,
    "size3": -0.738,
    "size4": -1.036,
    "size5": -1.123,
    "size6": -1.210,
    "size7": -1.452,
    "eduprimary": 0.141,
    "edusecondary": 0.168,
    "eduoversecondary": 0.416,
    "memberabroad": 0.244,
    "paidemp": 0.164,
    "selfemp": 0.178,
    "gas": 0.371,
    "paraffin": 0.122,
    "stove": 0.192,
    "radio": 0.141,
    "sheep": 0.215,
    "cattle": 0.099,
    "horses": 0.144,
    "poultry": 0.176,
}

POVERTY_MEAN = 3.866
POVERTY_SCALE = 1000 / 3.792
POVERTY_CUTOFF = 361

# ---------------------------------------------------------------------------
# Fuel Codes
# ---------------------------------------------------------------------------

FUEL_GAS = 2
FUEL_PARAFFIN = 3

# ---------------------------------------------------------------------------
# Activity Codes
# ---------------------------------------------------------------------------

PAID_EMPLOYEE_CODES = (1, 2)
SELF_EMPLOYEE_CODES = (4, 5)
STUDENT_CODE = 7

# ---------------------------------------------------------------------------
# Living Elsewhere Codes
# ---------------------------------------------------------------------------

LIVING_IN_HOUSE = 1
LIVING_ELSEWHERE = (2, 3, 4, 5)
LIVING_ABROAD = (3, 4)

# ---------------------------------------------------------------------------
# Excel Formatting
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

DATA_FONT = Font(name="Arial", size=10)

POOR_FILL = PatternFill("solid", start_color="FFCCCC")
BETTER_FILL = PatternFill("solid", start_color="CCFFCC")
MISSING_FILL = PatternFill("solid", start_color="FFF2CC")

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMN_GROUPS = {
    "Identification": ["Household_Id"],
    "Household Size": ["HH_Size", "In_House_Count", "Student_Count"],
    "Education & Work": [
        "Highest_LOE",
        "Member_Abroad",
        "Paid_Employee",
        "Self_Employee",
    ],
    "Assets": [
        "Heating_Fuel_Code",
        "var_gas",
        "var_paraffin",
        "var_stove",
        "var_radio",
        "var_sheep",
        "var_cattle",
        "var_horses",
        "var_poultry",
    ],
    "Size Dummies": [
        "var_size2",
        "var_size3",
        "var_size4",
        "var_size5",
        "var_size6",
        "var_size7",
    ],
    "Education Dummies": [
        "var_eduprimary",
        "var_edusecondary",
        "var_eduoversecondary",
    ],
    "PMT Output": [
        "PMT_Score",
        "Poverty_Level",
        "Parameter_Status",
        "Missing",
        "Missing_Reason",
        "Used_Values",
    ],
}

GROUP_COLORS = {
    "Identification": "BDD7EE",
    "Household Size": "DDEBF7",
    "Education & Work": "E2EFDA",
    "Assets": "FFF2CC",
    "Size Dummies": "FCE4D6",
    "Education Dummies": "EAD1DC",
    "PMT Output": "D9E1F2",
}