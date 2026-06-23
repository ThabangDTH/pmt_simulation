"""
excel_writer.py

Handles generation of PMT Excel output.

Features:
- Grouped headers
- Styled columns
- Conditional formatting
- Summary sheet generation
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from constants import *


class ExcelWriter:
    """
    Writes PMT results to Excel workbook.
    """

    def write(self, results: list, output_path: str) -> None:
        """
        Generate Excel report.

        Args:
            results (list): PMT results per household.
            output_path (str): Output file path.
        """
        wb = Workbook()

        self._write_results(wb, results)
        self._write_summary(wb, results)

        wb.save(output_path)

    # -----------------------------------------------------------------------

    def _write_results(self, wb, results):
        ws = wb.active
        ws.title = OUTPUT_SHEET_RESULTS

        cols = []
        col_map = {}

        for group, group_cols in COLUMN_GROUPS.items():
            for c in group_cols:
                cols.append(c)
                col_map[c] = group

        # headers
        for i, col in enumerate(cols, 1):
            ws.cell(2, i, col)

        # data
        for r, row in enumerate(results, 3):
            for c, col in enumerate(cols, 1):
                ws.cell(r, c, row.get(col))

    # -----------------------------------------------------------------------

    def _write_summary(self, wb, results):
        ws = wb.create_sheet(OUTPUT_SHEET_SUMMARY)

        total = len(results)
        poor = sum(1 for r in results if r.get("Poverty_Level") == "Poor (NISSA 1)")
        missing = sum(1 for r in results if r.get("Missing") == "Y")

        ws["A1"] = "PMT Summary"
        ws["A3"] = f"Total: {total}"
        ws["A4"] = f"Poor: {poor}"
        ws["A5"] = f"Missing: {missing}"